import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl.styles import Border, Side, Alignment

def seleccionar_archivo():
    """Abre una ventana para que seleccione un archivo CSV o Excel."""
    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename(
        title="Selecciona tu archivo de QGIS (CSV o Excel)",
        filetypes=(
            ("Archivos Soportados", "*.csv *.xlsx"),
            ("Archivos CSV", "*.csv"),
            ("Archivos Excel", "*.xlsx"),
            ("Todos los archivos", "*.*")
        )
    )
    return filepath

def aplicar_formato_final_excel(writer, df):
    """
    Aplica el formato final: enmarcado grueso, bordes de bloque,
    bordes interiores finos, y combina y centra celdas.
    """
    workbook = writer.book
    worksheet = writer.sheets['Vertices']
    
    thin_side = Side(style='thin')
    thick_side = Side(style='thick')
    alineacion_centrada = Alignment(horizontal='center', vertical='center', wrap_text=True)

    max_row = worksheet.max_row
    max_col = worksheet.max_column

    for c in range(1, max_col + 1):
        cell = worksheet.cell(row=1, column=c)
        cell.border = Border(top=thick_side, bottom=thick_side, 
                               left=thick_side if c == 1 else thin_side, 
                               right=thick_side if c == max_col else thin_side)

    start_row = 2
    for _, group in df.groupby('Nombre del sitio', sort=False):
        num_filas_grupo = len(group)
        end_row = start_row + num_filas_grupo - 1

        for r in range(start_row, end_row + 1):
            for c in range(1, max_col + 1):
                cell = worksheet.cell(row=r, column=c)
                borde_superior = thick_side if r == start_row else thin_side
                borde_inferior = thick_side if r == end_row or r == max_row else thin_side
                borde_izquierdo = thick_side if c == 1 else thin_side
                borde_derecho = thick_side if c == max_col else thin_side
                cell.border = Border(top=borde_superior, bottom=borde_inferior, 
                                       left=borde_izquierdo, right=borde_derecho)
        
        if num_filas_grupo > 1:
            worksheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        
        cell_a_centrar = worksheet.cell(row=start_row, column=1)
        cell_a_centrar.alignment = alineacion_centrada
        
        start_row = end_row + 1
        
    print("Formato de bordes y celdas combinadas aplicado.")

def procesar_archivo_de_datos(ruta_archivo):
    """
    Función principal que pide datos, lee, procesa y exporta el archivo con formato.
    """
    if not ruta_archivo:
        print("No se seleccionó ningún archivo. Saliendo del programa.")
        return

    huso_ingresado = input("Ahora, ingresa el valor para la columna 'Huso' (ej: 19S): ")
    if not huso_ingresado:
        huso_ingresado = "Sin especificar"

    try:
        print(f"\nLeyendo el archivo: {os.path.basename(ruta_archivo)}")
        
        if ruta_archivo.lower().endswith('.csv'):
            df = pd.read_csv(ruta_archivo)
        elif ruta_archivo.lower().endswith('.xlsx'):
            df = pd.read_excel(ruta_archivo)
        else:
            print("Error: Formato de archivo no soportado.")
            return
        
        print("Archivo leído correctamente.")

        df['name'] = df['UniqueID'].str.split('_').str[0]

        df_base = df[['name', 'ID_Vertice', 'Coordenada_X', 'Coordenada_Y']]
        df_base['Huso'] = huso_ingresado
        mapeo_nombres = {'name': 'Nombre del sitio', 'ID_Vertice': 'Vértice', 'Coordenada_X': 'UTM E', 'Coordenada_Y': 'UTM N'}
        df_final = df_base.rename(columns=mapeo_nombres)
        orden_final = ['Nombre del sitio', 'Vértice', 'UTM E', 'UTM N', 'Huso']
        df_final = df_final[orden_final]
        
        print("Datos procesados y ordenados.")

        conteo_de_sitios = df_final['Nombre del sitio'].nunique()
        print(f"\n--- Resumen del Proceso ---")
        print(f"Se han procesado {conteo_de_sitios} sitios únicos.")
        print(f"---------------------------\n")

        directorio = os.path.dirname(ruta_archivo)
        nombre_base = os.path.splitext(os.path.basename(ruta_archivo))[0]
        ruta_excel_salida = os.path.join(directorio, f"{nombre_base}_procesado.xlsx")

        with pd.ExcelWriter(ruta_excel_salida, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name='Vertices')
            aplicar_formato_final_excel(writer, df_final)

        print(f"¡Proceso completado! ✨")
        print(f"Tu reporte final ha sido guardado en: {ruta_excel_salida}")

    except Exception as e:
        print(f"\nOcurrió un error inesperado: {e}")

if __name__ == "__main__":
    ruta_del_archivo = seleccionar_archivo()
    procesar_archivo_de_datos(ruta_del_archivo)